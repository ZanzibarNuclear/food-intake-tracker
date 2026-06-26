from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


def cell_value(cell: Cell) -> Any:
    value = cell.value
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def non_empty_rows(sheet, max_rows: int = 80, max_cols: int = 24) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row, max_rows),
        min_col=1,
        max_col=min(sheet.max_column, max_cols),
    ):
        values = [cell_value(cell) for cell in row]
        if any(value not in (None, "") for value in values):
            rows.append(values)
    return rows


def formula_cells(sheet, limit: int = 200) -> list[dict[str, Any]]:
    formulas: list[dict[str, Any]] = []
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formulas.append({"cell": cell.coordinate, "formula": value})
                if len(formulas) >= limit:
                    return formulas
    return formulas


def merged_ranges(sheet) -> list[str]:
    return [str(merged) for merged in sheet.merged_cells.ranges]


def dimensions(sheet) -> dict[str, Any]:
    widths = {
        get_column_letter(index): sheet.column_dimensions[get_column_letter(index)].width
        for index in range(1, min(sheet.max_column, 24) + 1)
        if sheet.column_dimensions[get_column_letter(index)].width is not None
    }
    return {
        "maxRow": sheet.max_row,
        "maxColumn": sheet.max_column,
        "columnWidths": widths,
        "freezePanes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
    }


def workbook_summary(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    return {
        "workbook": path.name,
        "sheets": [
            {
                "name": sheet.title,
                "dimensions": dimensions(sheet),
                "mergedRanges": merged_ranges(sheet),
                "tables": list(sheet.tables.keys()),
                "sampleRows": non_empty_rows(sheet),
                "formulas": formula_cells(sheet),
            }
            for sheet in workbook.worksheets
        ],
    }


def main() -> None:
    if len(sys.argv) != 3:
        print("Usage: extract_workbook.py <input.xlsx> <output.json>", file=sys.stderr)
        raise SystemExit(2)

    input_path = Path(sys.argv[1]).expanduser()
    output_path = Path(sys.argv[2]).expanduser()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = workbook_summary(input_path)
    output_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


if __name__ == "__main__":
    main()
