from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def value_to_json(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def compact_text(value: Any) -> str:
    return str(value or "").strip()


def extract_foods(sheet) -> list[dict[str, Any]]:
    foods: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = compact_text(row[0])
        if not name:
            continue
        foods.append(
            {
                "name": name,
                "servingDescription": compact_text(row[1]),
                "calories": float(row[2] or 0),
                "proteinGrams": float(row[3] or 0),
                "nutritionScore": float(row[4] or 0),
                "satietyScore": float(row[5]) if row[5] is not None else None,
                "notes": compact_text(row[6]) or None,
            }
        )
    return foods


def extract_meals(sheet) -> list[dict[str, Any]]:
    meals: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        date = value_to_json(row[0])
        meal = compact_text(row[1])
        food_name = compact_text(row[2])
        quantity = row[3]
        if not date or not meal or not food_name:
            continue
        meals.append(
            {
                "date": date,
                "meal": meal,
                "foodName": food_name,
                "quantity": float(quantity or 1),
                "notes": compact_text(row[7]) or None,
            }
        )
    return meals


def extract_weights(sheet) -> list[dict[str, Any]]:
    weights: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        date = value_to_json(row[0])
        weight = row[1]
        if not date or weight is None:
            continue
        weights.append(
            {
                "date": date,
                "weight": float(weight),
                "goalWeight": float(row[2]) if row[2] is not None else None,
                "notes": compact_text(row[3]) or None,
            }
        )
    return weights


def main() -> None:
    if len(sys.argv) != 3:
        print("Usage: extract_seed.py <input.xlsx> <output.json>", file=sys.stderr)
        raise SystemExit(2)

    input_path = Path(sys.argv[1]).expanduser()
    output_path = Path(sys.argv[2]).expanduser()
    workbook = load_workbook(input_path, data_only=False)
    seed = {
        "settings": {
            "dailyCalorieTarget": 2000,
            "proteinTargetGrams": 100,
            "nutritionScoreTarget": 7,
            "goalWeight": 170,
        },
        "foods": extract_foods(workbook["Foods"]),
        "meals": extract_meals(workbook["Meal Log"]),
        "weights": extract_weights(workbook["Weight Log"]),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(seed, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
